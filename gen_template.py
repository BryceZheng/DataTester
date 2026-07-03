#!/usr/bin/env python3
"""生成 DataTester 判断规则配置模板到桌面。"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "判断规则"

TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FILL   = PatternFill("solid", fgColor="BDD7EE")
INST_FILL  = PatternFill("solid", fgColor="DEEAF1")
WHITE      = PatternFill("solid", fgColor="FFFFFF")
GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
HDR_FONT   = Font(name="微软雅黑", bold=True, size=10, color="1F4E79")
INST_FONT  = Font(name="微软雅黑", italic=True, size=9, color="595959")
BODY_FONT  = Font(name="微软雅黑", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="B4C6E7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c

# 标题行
ws.merge_cells("A1:F1")
set_cell(ws, 1, 1, "DataTester 判断规则模板", font=TITLE_FONT, fill=TITLE_FILL, align=CENTER)
ws.row_dimensions[1].height = 32

# 说明行
ws.merge_cells("A2:F2")
set_cell(ws, 2, 1,
    "填写本次的判断规则后，将此文件路径提供给 DataTester（输入/输出文件的分析由 Claude 自动完成）",
    font=INST_FONT, fill=INST_FILL, align=CENTER)
ws.row_dimensions[2].height = 18

ws.merge_cells("A3:F3")
set_cell(ws, 3, 1,
    "  说明：「适用字段」填输入 Excel 的列标识，如 A列、C列、D列；规则冲突时优先级数字越小越优先",
    font=INST_FONT, fill=INST_FILL, align=LEFT)
ws.row_dimensions[3].height = 18

# 表头行
HEADERS = [
    "适用字段\n(如 A列、C列)",
    "判断条件\n(文字描述，如：值为空、包含'已签'、大于100)",
    "输出字段\n(对应输出列名)",
    "输出值\n(满足条件时填写的内容)",
    "优先级\n(数字越小越优先)",
    "备注",
]
for i, h in enumerate(HEADERS, 1):
    set_cell(ws, 4, i, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)
ws.row_dimensions[4].height = 32

# 空白数据行（20行）
for i in range(20):
    r = 5 + i
    fill = GRAY_FILL if i % 2 == 1 else WHITE
    for c in range(1, 7):
        set_cell(ws, r, c, None, font=BODY_FONT, fill=fill, align=LEFT, border=BORDER)
    ws.row_dimensions[r].height = 20

# 列宽
col_widths = [16, 40, 18, 28, 12, 16]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

output = os.path.expanduser("~/Desktop/DataTester判断规则模板.xlsx")
wb.save(output)
print(f"✅ 模板已生成：{output}")
