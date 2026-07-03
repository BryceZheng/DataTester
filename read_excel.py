#!/usr/bin/env python3
"""
read_excel.py — DataTester helper
Usage:
  python3 read_excel.py <file_path> <start_row> <end_row> [--header-row N]
  python3 read_excel.py <file_path> count [--header-row N]

  --header-row N  treat row N as the header row (default: 1)
                  useful for files with title/instruction rows above the real headers

Output: JSON to stdout with keys:
  - headers: list of column header strings (from the designated header row)
  - rows: list of dicts (column name -> cell value)
  - total_rows: total data rows (rows after header row)
  - merge_info: list of merged cell range strings (e.g. "A1:C3")
"""
import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter


def cell_value(cell):
    v = cell.value
    if v is None:
        return None
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return v


def main():
    # Parse --header-row N from anywhere in argv
    args = sys.argv[1:]
    header_row = 1
    if "--header-row" in args:
        idx = args.index("--header-row")
        header_row = int(args[idx + 1])
        args = args[:idx] + args[idx + 2:]

    if len(args) < 2:
        print("Usage: read_excel.py <file_path> <start_row|count> [end_row] [--header-row N]", file=sys.stderr)
        sys.exit(1)

    file_path = args[0]
    mode = args[1]

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    if mode == "count":
        data_rows = max_row - header_row if max_row > header_row else 0
        print(json.dumps({"total_rows": data_rows, "total_with_header": max_row, "header_row": header_row}, ensure_ascii=False))
        return

    start_row = int(mode)
    end_row = int(args[2]) if len(args) > 2 else max_row

    # Clamp to actual sheet bounds
    start_row = max(1, start_row)
    end_row = min(max_row, end_row)

    # Read headers from designated header row
    headers = []
    for col in range(1, max_col + 1):
        h = ws.cell(row=header_row, column=col).value
        headers.append(str(h) if h is not None else f"Column_{get_column_letter(col)}")

    # Read requested rows
    rows = []
    for row_idx in range(start_row, end_row + 1):
        row_dict = {"_row": row_idx}
        for col_idx, header in enumerate(headers, start=1):
            row_dict[header] = cell_value(ws.cell(row=row_idx, column=col_idx))
        rows.append(row_dict)

    # Collect merge info
    merge_info = [str(r) for r in ws.merged_cells.ranges]

    output = {
        "file": file_path,
        "sheet": ws.title,
        "header_row": header_row,
        "headers": headers,
        "start_row": start_row,
        "end_row": end_row,
        "rows": rows,
        "total_rows": max_row - header_row,
        "merge_info": merge_info,
    }

    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
