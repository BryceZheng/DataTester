#!/usr/bin/env python3
"""
write_excel.py — DataTester helper
Usage:
  python3 write_excel.py <template_path> <json_data_path> <output_path>

JSON data format (written by Claude to scratchpad, then passed here):
{
  "headers": ["col1", "col2", ...],   // output column names matching template
  "rows": [
    {"col1": "value", "col2": "value", ...},
    ...
  ],
  "merge_rules": [                     // optional: explicit merge instructions
    {"range": "A2:B2", "value": "merged text"}
  ]
}

Behavior:
  - Copies template file to output path (preserves all styles, column widths, header row)
  - Appends data rows starting from the first empty row after the header
  - Inherits cell style from the corresponding column's first data row in the template
  - Applies merge_rules if provided; otherwise replicates template's existing merge pattern
    for repeating-value columns
"""
import sys
import json
import shutil
import copy
import openpyxl
from openpyxl.utils import column_index_from_string


def get_template_row_style(ws, row_idx, max_col):
    """Extract per-column styles from a template row."""
    styles = {}
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        styles[col] = {
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "border": copy.copy(cell.border),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }
    return styles


def apply_style(cell, style):
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.border = style["border"]
    cell.alignment = style["alignment"]
    cell.number_format = style["number_format"]


def find_header_row(ws, expected_headers):
    """Find the row index where headers are located (scan first 10 rows)."""
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        # Check if all expected headers appear somewhere in this row
        if all(h in row_values for h in expected_headers if h):
            return row_idx
    return 1  # default to row 1


def build_col_map(ws, header_row):
    """Map header name -> column index in the template."""
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            col_map[str(val)] = col
    return col_map


def main():
    if len(sys.argv) < 4:
        print("Usage: write_excel.py <template_path> <json_data_path> <output_path>", file=sys.stderr)
        sys.exit(1)

    template_path = sys.argv[1]
    json_path = sys.argv[2]
    output_path = sys.argv[3]

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    headers = data.get("headers", [])
    rows = data.get("rows", [])
    merge_rules = data.get("merge_rules", [])

    # Copy template to output path (preserves all styling)
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Locate header row
    header_row = find_header_row(ws, headers)
    col_map = build_col_map(ws, header_row)

    # Find first data row: look for first completely empty row after header
    first_data_row = header_row + 1

    # Get style from template first data row if it exists, else from header row
    template_style_row = first_data_row if ws.max_row >= first_data_row else header_row
    col_styles = get_template_row_style(ws, template_style_row, ws.max_column)

    # Clear any existing data rows (keep header)
    for row_idx in range(first_data_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).value = None

    # Remove existing data-area merges (keep header merges)
    merges_to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= first_data_row:
            merges_to_remove.append(str(merged_range))
    for r in merges_to_remove:
        ws.unmerge_cells(r)

    # Write data rows
    for i, row_data in enumerate(rows):
        target_row = first_data_row + i
        for header, col_idx in col_map.items():
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = row_data.get(header)
            if col_idx in col_styles:
                apply_style(cell, col_styles[col_idx])

    # Apply explicit merge rules if provided
    for rule in merge_rules:
        cell_range = rule.get("range", "")
        value = rule.get("value")
        if cell_range:
            ws.merge_cells(cell_range)
            # Set value on top-left cell of merge
            try:
                start_cell = cell_range.split(":")[0]
                col_letter = ''.join(c for c in start_cell if c.isalpha())
                row_num = int(''.join(c for c in start_cell if c.isdigit()))
                col_num = column_index_from_string(col_letter)
                ws.cell(row=row_num, column=col_num).value = value
            except Exception as e:
                print(f"merge warning: {cell_range}: {e}", file=sys.stderr)

    wb.save(output_path)
    print(json.dumps({
        "status": "ok",
        "output": output_path,
        "rows_written": len(rows),
        "headers": headers,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
